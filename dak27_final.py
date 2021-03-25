import pandas as pd
import psycopg2
import xlsxwriter
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import datetime
from pytrends.request import TrendReq


def connect():
    ##Connect to the PostgreSQL database server
    conn, cur = None, None
    try:
        # connect to the PostgreSQL server
        print('Connecting to the PostgreSQL database...')
        conn = psycopg2.connect(database="dak27", user="postgres", password="huy6257001", host="127.0.0.1", port="5433")
        # create a cursor
        cur = conn.cursor()
    except (Exception, psycopg2.DatabaseError) as error:
        print("Error while excuting SQL" + error)

    return conn, cur


def chunkIt(seq, num):
    ## Split List with the maximun element
    const = num
    out = []
    last = 0.0
    if len(seq) <= const:
        return out.append(seq[0:(len(seq) - 1)])
    else:
        while last < len(seq):
            out.append(seq[int(last):int(last + const)])
            last += const
        return out


def formatString(term):
    character = "\'"
    if character in term:
        term = term.replace(character, "")
    return term


def insertDB(keyword, date, value, trend_type):
    con, cur = connect()
    print("Database opened successfully")
    value = "VALUES ('{}', '{}', '{}', '{}')".format(formatString(keyword), date, value, trend_type)
    query = "INSERT INTO vn_trending (keyword,date,value,trend_type) ({})".format(value)
    cur.execute(query)
    con.commit()
    print("Record inserted successfully")
    con.close()


def build_sql(year=2020, limit=10):
    sql = """SELECT row_number() over (ORDER BY A.sum_val DESC, A.keyword, B.monthly) as stt , 
                    A.keyword, A.sum_val, B.monthly, B.max_val
            FROM
                (	SELECT keyword, sum(VALUE::INT) sum_val
                    FROM vn_trending
                    WHERE EXTRACT(YEAR FROM DATE) = %s
                    GROUP BY keyword
                    ORDER BY sum(VALUE::INT) DESC
                    LIMIT %s
                ) A
            JOIN 
                (
                    SELECT DISTINCT A1.keyword, A2.monthly, A1.max_val
                    FROM
                    (	SELECT keyword, max(sum_val) max_val
                        FROM
                        (
                            SELECT DISTINCT keyword, sum(VALUE::INT) sum_val, to_char(date::date,'mm/yyyy') monthly
                            FROM vn_trending
                            WHERE EXTRACT(YEAR FROM DATE) = %s
                            GROUP BY keyword,to_char(date::date,'mm/yyyy') 
                            ORDER BY keyword, sum(VALUE::INT) DESC
                        ) A3 
                        GROUP BY keyword
                    ) A1
                    JOIN
                    (		SELECT DISTINCT keyword, sum(VALUE::INT) sum_val, to_char(date::date,'mm/yyyy') monthly
                            FROM vn_trending
                            WHERE EXTRACT(YEAR FROM DATE) = %s
                            GROUP BY keyword,to_char(date::date,'mm/yyyy') 
                            ORDER BY keyword, sum(VALUE::INT) DESC
                        ) A2 ON A1.keyword = A2.keyword and A1.max_val = A2.sum_val
                ) B ON A.keyword = B.keyword
            ORDER BY A.sum_val DESC, A.keyword, B.monthly;""" % (year, limit, year, year)

    return sql


## Menu ##
def show_action():
    print("Danh sách Menu")
    for i in listMenu:
        print(i)


def action():
    loop = True
    while loop:
        show_action()
        player_action = input("Nhập chỉ mục: ")
        if player_action == '1':
            while True:
                excel_name = input("Nhập tên excel: ")
                try:
                    df = pd.read_excel(r'E:\Documents\Documents\DocProjects\Talent5\{0}.xlsx'.format(excel_name))
                except Exception as e:
                    print(e)
                    continue
                break

            while True:
                try:
                    print("Nhập thời gian bắt đầu")
                    start_year = int(input('Nhập năm: '))
                    start_month = int(input('Nhập tháng: '))
                    start_day = int(input('Nhập ngày: '))
                    start_date = datetime.date(start_year, start_month, start_day)

                    print("Nhập thời gian kết thúc")
                    end_year = int(input('Nhập năm: '))
                    end_month = int(input('Nhập tháng: '))
                    end_day = int(input('Nhập ngày: '))
                    end_date = datetime.date(end_year, end_month, end_day)
                except ValueError:
                    print("Nhập sai, nhập lại")
                    continue
                if end_date >= start_date:
                    break
                else:
                    print("Nhập sai, Thời gian bắt đầu phải sau thời gian kết thúc")

            home_act1(df, start_date, end_date)  # call the respective action function here

        elif player_action == '2':

            while True:
                try:
                    year = int(input('Nhập năm: '))
                except ValueError:
                    print("Nhập sai, nhập lại")
                    continue
                if len(str(year)) == 4:
                    break
                else:
                    print("Nhập sai, nhập lại")

            home_act2(year)

        elif player_action == '3':
            home_act3()

        elif player_action == '4':
            home_act4()

        elif player_action == '5':
            home_act5()

        elif player_action == '6':
            home_act6()

        elif player_action == '99':
            exit(1)
        else:
            print("Không có chỉ mục phù hợp, xin chọn lại!!!")


def home_act1(df, start_date, end_date):
    colnames = list(df.columns)
    arrDf=[]
    for x in colnames:
        df2 = df[x].values.tolist()
        cleanedList = [x for x in df2 if str(x) != 'nan']
        # module 2 - Send request into Google Trend
        # looping for checking batch, a batch only contain 5 keyworks
        keyw = []
        t = 0
        for y in chunkIt(cleanedList, 5):
            pytrend = TrendReq('vi-VI')
            pytrend.build_payload(
                y,
                timeframe='{0} {1}'.format(start_date, end_date),
                geo='VN',
                gprop=''
            )

            data = pytrend.interest_over_time()

            if t != 0:
                # Conver data to array at the same time define the column name
                # iterating the dataFrame to define the column name
                result = pd.concat([result, data], axis=1)
            else:
                result = pytrend.interest_over_time()
            t = t + 1
        del result["isPartial"]
        print(result)
        arrDf.append(result)

    for arr in arrDf:
        # print(arr)
        for i in range(len(arr)):
            for j in range(len(arr.columns)):
                insertDB(arr.columns[j], arr.index[i], arr.iloc[i, j], x)
                print("Record inserted successfully")

    # #Insert Database
    # for i in range(len(result)):
    #     for j in range(len(result.columns)):
    #         print("Record inserted successfully")
    #         insertDB(result.columns[j], result.index[i], result.iloc[i, j], x)


def home_act2(year):
    sql = """
                SELECT 
                row_number() over (ORDER BY A.rateSearch DESC, A.keyword, B.monthly) as stt , 
                        A.keyword, A.rateSearch, B.monthly, B.max_val
                FROM
                    (	SELECT keyword, sum(VALUE::INT) rateSearch
                        FROM vn_trending
                        WHERE EXTRACT(YEAR FROM DATE) = %s
                        GROUP BY keyword
                        ORDER BY sum(VALUE::INT) DESC
                        LIMIT 10
                    ) A
                JOIN 
                    (
                        SELECT DISTINCT A1.keyword, A2.monthly, A1.max_val
                        FROM
                        (	SELECT keyword, max(rateSearch) max_val
                            FROM
                            (
                                SELECT DISTINCT keyword, sum(VALUE::INT) rateSearch, to_char(date::date,'mm/yyyy') monthly
                                FROM vn_trending
                                WHERE EXTRACT(YEAR FROM DATE) = %s
                                GROUP BY keyword,to_char(date::date,'mm/yyyy') 
                                ORDER BY keyword, sum(VALUE::INT) DESC
                            ) A3 
                            GROUP BY keyword
                        ) A1
                        JOIN
                        (		SELECT DISTINCT keyword, sum(VALUE::INT) rateSearch, to_char(date::date,'mm/yyyy') monthly
                                FROM vn_trending
                                WHERE EXTRACT(YEAR FROM DATE) = %s
                                GROUP BY keyword,to_char(date::date,'mm/yyyy') 
                                ORDER BY keyword, sum(VALUE::INT) DESC
                            ) A2 ON A1.keyword = A2.keyword and A1.max_val = A2.rateSearch
                    ) B ON A.keyword = B.keyword
                ORDER BY A.rateSearch DESC, A.keyword, B.monthly;
        """ % (year, year, year)
    conn, cur = connect()
    cur.execute(sql)
    rd = cur.fetchall()
    conn.close()
    cur.close()

    df = pd.DataFrame(rd, columns=['STT','Keyword', 'Số lượt tìm kiếm', 'Thời gian', 'Số lượt tìm kiếm trong tháng'])
    # df.index += 1

    filename = 'vn_trending_top_ten.xlsx'
    writer = pd.ExcelWriter(filename)
    df.to_excel(writer, index=False)
    writer.save()

    wb = load_workbook(filename)
    # Select First Worksheet
    ws = wb.worksheets[0]

    ws.insert_rows(0, 2)
    ws.cell(1, 1, 'DANH SÁCH TỪ KHÓA TÌM KIẾM TẠI VIỆT NAM')
    ws.cell(2, 1, 'Năm: 2020')

    print("Insert into Excel successfully")
    wb.save(filename)


def home_act3():
    sql = """SELECT 
            A.keyword, A.rateSearch, A.monthly, A.trend_type
            FROM
            (	SELECT trend_type, keyword, sum(VALUE::INT) rateSearch, to_char(date::date,'mm/yyyy') monthly
                FROM vn_trending
                WHERE EXTRACT(YEAR FROM DATE) = '2020'
                GROUP BY keyword, trend_type, to_char(date::date,'mm/yyyy')
                ORDER BY sum(VALUE::INT) DESC
            ) A;
            """
    conn, cur = connect()
    cur.execute(sql)
    rd = cur.fetchall()

    conn.close()
    cur.close()

    filename = 'vn_trending_search_keyword_2020.xlsx'
    df = pd.DataFrame(rd, columns=['keyword', 'rateSearch', 'monthly', 'trend_type'])
    # print(df) cải thiện để thêm cột STT
    df2 = df.pivot_table(index="keyword", columns='monthly', values='rateSearch')
    writer = pd.ExcelWriter(filename)
    df2.to_excel(writer)
    writer.save()

    wb = load_workbook(filename)
    # Select First Worksheet
    ws = wb.worksheets[0]

    ws.insert_rows(0, 2)
    ws.cell(1, 1, 'TỪ KHÓA TÌM KIẾM NHIỀU NHẤT TẠI VIỆT NAM')
    ws.cell(2, 1, 'Năm: 2020')

    print("Insert into Excel successfully")
    wb.save(filename)


def home_act4():
    conn, cur = connect()
    sql = build_sql('2020', 5)
    cur.execute(sql)
    rd = cur.fetchall()
    df = pd.DataFrame(rd, columns=['stt', 'keyword', 'sum_val', 'monthly', 'max_val'])
    # df.plot(x='keyword', y='sum_val', kind='line')
    # plt.title('TỪ KHÓA TÌM KIẾM NHIỀU NHẤT TẠI VIỆT NAM 2020')
    # plt.show()
    if len(df):
        image = df.plot(x='keyword', y='sum_val', legend=False, kind='line',
                        title='TỪ KHÓA TÌM KIẾM NHIỀU NHẤT TẠI VIỆT NAM 2020', figsize=(10, 3))
        fig = image.get_figure()
        fig.savefig('top_search_key_2020.png')


def home_act5():
    conn, cur = connect()
    sql = build_sql('2019', 5)
    cur.execute(sql)
    rd = cur.fetchall()
    df = pd.DataFrame(rd, columns=['stt', 'keyword', 'sum_val', 'monthly', 'max_val'])
    # df.plot(x='keyword', y='sum_val', kind='line')
    # plt.title('TỪ KHÓA TÌM KIẾM NHIỀU NHẤT TẠI VIỆT NAM 2020')
    # plt.show()
    if len(df):
        image = df.plot(x='keyword', y='sum_val', legend=False, kind='bar',
                        title='TỪ KHÓA TÌM KIẾM NHIỀU NHẤT TẠI VIỆT NAM 2019', figsize=(10, 3))
        fig = image.get_figure()
        fig.savefig('top_search_key_2019.png')
    else:
        print('Null Value')


def home_act6():
    conn, cur = connect()

    sql = build_sql('2020', 5)
    cur.execute(sql)
    rd = cur.fetchall()
    df2020 = pd.DataFrame(rd, columns=['STT', 'Keyword', 'Số lần tìm kiếm', 'Tháng tìm kiếm nhiều nhất', 'Lượt tìm kiếm nhiều nhất trong tháng'])
    df2020.index += 1
    df2020.name = "Năm 2020"
    del df2020["Lượt tìm kiếm nhiều nhất trong tháng"]

    sql = build_sql('2019', 5)
    cur.execute(sql)
    rd = cur.fetchall()
    df2019 = pd.DataFrame(rd, columns=['STT', 'Keyword', 'Số lần tìm kiếm', 'Tháng tìm kiếm nhiều nhất', 'Lượt tìm kiếm nhiều nhất trong tháng'])
    df2019.index += 1
    df2019.name = "Năm 2019"
    del df2019["Lượt tìm kiếm nhiều nhất trong tháng"]
    del df2019["STT"]
    # df = pd.concat([df2020, df2019], join='outer')
    # print(df)
    print(df2019)
    print(df2020)

    filename = r'vn_trending_top5_search_keyword_2019_2020.xlsx'
    writer = pd.ExcelWriter(filename, engine='xlsxwriter')
    df2020.to_excel(writer, startrow=0 , startcol=0,index=False)
    df2019.to_excel(writer, startrow=0, startcol=4, index=False)

    writer.save()

    wb = load_workbook(filename)
    # sheet = wb.active
    # Select First Worksheet
    ws = wb.worksheets[0]

    ws.insert_rows(0, 2)
    #Merge Cell Ask Mentor
    # ws.merge_cells('A1:A6')

    ws.cell(1, 1, 'DANH SÁCH TỪ KHÓA TÌM KIẾM TẠI VIỆT NAM')
    ws.cell(2, 1, 'Năm: 2020')
    ws.cell(2, 5, 'Năm: 2019')

    print("Insert into Excel successfully")
    wb.save(filename)


## Main Fucntion ##
s1 = '1. Lấy dữ liệu trending từ file. ';
s2 = '2. Xuất báo cáo top 10 trending';
s3 = '3. Xuất báo cáo search keyword in 2020';
s4 = '4. Vẽ biểu đồ line chart top 5 trending các từ khóa tìm kiếm nhiều nhất 2020';
s5 = '5. Vẽ biểu đồ bar chart top 5 trending các từ khóa tìm kiếm nhiều nhất 2019';
s6 = '6. Thống kê tìm kiếm top trending 5 từ khóa trong 2 năm 2020, 2019';
s99 = '99. Thoát'
listMenu = [s1, s2, s3, s4, s5, s6, s99]

action()
